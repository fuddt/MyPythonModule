# Excelに合わせてデータフレームを作成
import pandas as pd
import openpyxl
from itertools import islice


class ExcelView:
    """
    how to use:

    #Excelファイルを読み込む
    ev = ExcelView('sample.xlsx')

    -------------------------------------------
    #Excelの列名と行番号に準じた形式で表示したい
    ・行番号は1から始まる
    ・列名はAから始まる

    ev.view_sheet('Sheet1', format='excel')
    >>
            A	        B	        C	    D	
    1	RecordSpec	DataKubun	MakeDate	Year	
    2	SE	            A	    20050303	1987	
    3	SE	            A	    20060508	1987	
    4	SE	            A	    20050303	1987	
    
    -----------------------------------------------
    #pandasのDataFrameに準じた形式で表示したい
    ・行番号は0から始まる
    ・列名は1行目のデータに準じる
    ev.view_sheet('Sheet1', format='pandas')
    >>
        DataKubun	MakeDate	Year	MonthDay	
    0	A	        20050303	1987	0519	
    1	A	        20060508	1987	0602	
    2	A	        20050303	1987	0615

    -----------------------------------------------
    #openpyxlの公式ドキュメント推奨の方法で表示したい
    ev.view_sheet('Sheet1', format='reference')
    ・行は1列目のデータに準じる
    ・列は1行目のデータに準じる
    >>
        DataKubun	MakeDate	Year	MonthDay	
    SE	A	        20050303	1987	0519	    
    SE	A	        20060508	1987	0602	    


    """



    def __init__(self, filename):
        self.workbook = openpyxl.load_workbook(filename)
        self.sheetnames = self.workbook.sheetnames

    @staticmethod
    def _to_excel_column_name(n):
        """
        Excelの列名に変換する関数
        ASCIIコードでは、65は文字「A」を表します
        """
        if n < 26:
            return chr(n + 65)
        else:
            return chr(n // 26 + 64) + chr(n % 26 + 65)
    @staticmethod
    def _to_excel_index(n):
        # インデックスを作成（Excelの行番号）
        return list(range(1, n+1))
    
    def view_sheet(self, sheetname, format='reference'):
        sheet = self.workbook[sheetname]
        if format == 'excel':
            return self._view_sheet_as_excel(sheet)
        elif format == 'pandas':
            return self._view_sheet_as_pandas(sheet)
        elif format == 'reference':
            return self._view_sheet_as_reference(sheet)
        else:
            raise ValueError(f'format {format} is not supported')

    def _view_sheet_as_excel(self, sheet):
        """
        Excelの列名と行番号に準じた形式
        """
        df = pd.DataFrame(sheet.values)
        column_names = list(map(self._to_excel_column_name, range(df.shape[1])))
        df.columns = column_names
        # インデックスを作成（Excelの行番号）
        idx = self._to_excel_index(df.shape[0])
        df.index = idx
        return df

    def _view_sheet_as_pandas(self, sheet):
        """
        pandasのDataFrameに準じた形式
        """
        data = sheet.values
        cols = next(data)[1:]
        data = list(data)
        data = (islice(r, 1, None) for r in data)
        return  pd.DataFrame(data, columns=cols)
    
    def _view_sheet_as_reference(self, sheet):
        """
        openpyxlの公式ドキュメント推奨の方法
        """
        data = sheet.values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        return pd.DataFrame(data, index=idx, columns=cols)