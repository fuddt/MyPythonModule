from openpyxl import load_workbook
import os
# Excelファイルを読み込み
current_dir = os.path.dirname(__file__) 
wb = load_workbook(os.path.join(current_dir, 'example.xlsx'))
ws = wb.active  # アクティブなシートを選択
def get_last_row_in_column(ws, column):
    last_row = 0
    for cell in ws[column]:  # 特定の列だけをイテレート
        print(type(cell.value))
        if isinstance(cell.value, str):  # セルの値が文字列かどうかチェック
            last_row = cell.row  # 文字列がある場合、行番号を更新
    return last_row
# C列における文字列データの最終行を取得
last_row_in_C = get_last_row_in_column(ws, 'C')
print(f'C列の文字列データのある最終行は {last_row_in_C} 行目です。')
# ワークブックを閉じる
wb.close()
