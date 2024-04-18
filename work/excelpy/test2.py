from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy

def copy_sheet_with_styles(src_sheet, dst_sheet):
    for row in src_sheet:
        for cell in row:
            new_cell = dst_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

# 元のワークブックをロード
src_wb = load_workbook('template.xlsx')
src_sheet = src_wb.active

# 新しいワークブックを作成
dst_wb = Workbook()
dst_sheet = dst_wb.active

# シートのコピー
copy_sheet_with_styles(src_sheet, dst_sheet)

# 新しいワークブックを保存
dst_wb.save('new_file.xlsx')