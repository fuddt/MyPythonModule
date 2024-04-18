import openpyxl

# Excelファイルを読み込む
file_path = 'sample.xlsx'
try:
    workbook = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

sheet = workbook.active  # アクティブなシートを取得

# 書式をコピーする最終行を検出
max_row = sheet.max_row

# 書式をコピー
for col in 'ABC':  # A, B, C列に対して
    last_cell = sheet[f'{col}{max_row}']
    for i in range(max_row + 1, max_row + 21):
        new_cell = sheet[f'{col}{i}']
        if last_cell.has_style:  # 最終行のセルにスタイルがある場合
            new_cell.font = copy(last_cell.font)
            new_cell.border = copy(last_cell.border)
            new_cell.fill = copy(last_cell.fill)
            new_cell.number_format = copy(last_cell.number_format)
            new_cell.protection = copy(last_cell.protection)
            new_cell.alignment = copy(last_cell.alignment)

# データを追加
for i in range(max_row + 1, max_row + 21):
    sheet[f'B{i}'].value = f'データ {i}'

# ファイルを保存
workbook.save(file_path)