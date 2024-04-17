import os
import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import glob

class FileManager:
    """
    ディレクトリ内のファイル操作を管理するクラス。
    
    Attributes:
        parent_dir (str): ファイルが格納されている親ディレクトリのパス。
    """
    def __init__(self, parent_dir):
        """
        FileManagerクラスのコンストラクタ。
        
        Args:
            parent_dir (str): ファイルが格納されている親ディレクトリのパス。
        """
        self.parent_dir = parent_dir
    
    def get_files_in_folder(self, folder):
        """
        指定されたフォルダ内のファイル名をリストとして返す。
        
        Args:
            folder (str): ファイル名を取得したいサブフォルダ名。
        
        Returns:
            list: 指定フォルダ内のファイル名のリスト。
        """
        folder_path = os.path.join(self.parent_dir, folder)
        return os.listdir(folder_path)

class ExcelManager:
    """
    Excelファイルの操作を管理するクラス。
    
    Attributes:
        excel_file (str): 操作するExcelファイルのパス。
        wb (openpyxl.workbook.Workbook): openpyxlを通じて操作するワークブックオブジェクト。
    """
    def __init__(self, excel_file):
        """
        ExcelManagerクラスのコンストラクタ。
        
        Args:
            excel_file (str): 操作するExcelファイルのパス。
        """
        self.excel_file = excel_file
        if os.path.exists(self.excel_file):
            self.wb = openpyxl.load_workbook(self.excel_file)
        else:
            self.wb = openpyxl.Workbook()
    
    def add_files_to_sheet(self, folder, files):
        """
        指定されたフォルダ名のシートにファイル名を追加する。
        
        Args:
            folder (str): シート名として使用するフォルダ名。
            files (list): シートに追加するファイル名のリスト。
        """
        if folder in self.wb.sheetnames:
            sheet = self.wb[folder]
        else:
            sheet = self.wb.create_sheet(title=folder)
        
        sheet['B3'] = '日付'
        sheet['C3'] = 'ファイル名'
        
        for i, file_name in enumerate(files, start=4):
            sheet[f'C{i}'] = file_name
        
        if len(files) > 0:
            self._merge_and_add_date(sheet, 4, 4 + len(files) - 1)
            self._add_borders(sheet, 4, 4 + len(files) - 1)
        
    def _merge_and_add_date(self, sheet, start_row, end_row):
        """
        指定された範囲のセルを結合し、現在の日付を挿入する。
        
        Args:
            sheet (openpyxl.worksheet.worksheet.Worksheet): 操作対象のシート。
            start_row (int): 結合を開始する行番号。
            end_row (int): 結合を終了する行番号。
        """
        sheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        sheet['B4'] = datetime.datetime.now().date()
        sheet['B4'].alignment = Alignment(horizontal='center', vertical='center')
    
    def _add_borders(self, sheet, start_row, end_row):
        """
        指定された範囲のセルに罫線を追加する。
        
        Args:
            sheet (openpyxl.worksheet.worksheet.Worksheet): 操作対象のシート。
            start_row (int): 罫線を追加する開始行番号。
            end_row (int): 罫線を追加する終了行番号。
        """
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in range(start_row, end_row + 1):
            for col in ['B', 'C']:
                cell = sheet[f'{col}{row}']
                cell.border = thin_border
    
    def save(self):
        """
        ワークブックに行った変更をファイルに保存する。
        """
        self.wb.save(self.excel_file)

# パスの設定
excel_file = './your_file.xlsx'
parent_dir = './root'
folders = [os.path.basename(folder) for folder in glob.glob(f'{parent_dir}/*')]

# インスタンスの生成
file_manager = FileManager(parent_dir)
excel_manager = ExcelManager(excel_file)

# フォルダごとにファイルを処理
for folder in folders:
    files = file_manager.get_files_in_folder(folder)
    excel_manager.add_files_to_sheet(folder, files)

# Excelファイルの保存
excel_manager.save()
