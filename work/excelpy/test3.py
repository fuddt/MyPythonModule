import openpyxl

# Excelファイルを読み込む（ファイルが存在しない場合は新規作成する）
file_path = 'sample.xlsx'
try:
    workbook = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

sheet = workbook.active  # アクティブなシートを取得

# B列の最終行を検出
max_row = sheet.max_row
if all(sheet[f'B{row}'].value is None for row in range(1, max_row + 1)):
    start_row = 1  # B列が全て空の場合、最初から開始
else:
    start_row = max_row + 1  # データのある最後の行の次から開始

# B列に20行データを追加
for i in range(start_row, start_row + 20):
    cell = sheet[f'B{i}']
    cell.value = f'データ {i}'  # ここに具体的なデータを設定する

# ファイルを保存
workbook.save(file_path)